import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

def create_base_file(actual_date):
    book = openpyxl.Workbook()
    sh = book.active
    sh.cell(1,2).value = "Place"
    sh.cell(1,3).value = "Points"
    sh.cell(1,4).value = "Team"
    sh.cell(1,4).alignment = Alignment(horizontal='center', vertical='center')
    sh.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8)
    sh.cell(1,9).value = "Url"

    sh.column_dimensions["A"].width = 10
    sh.column_dimensions["D"].width = 25
    sh.column_dimensions["E"].width = 25
    sh.column_dimensions["F"].width = 25
    sh.column_dimensions["G"].width = 25
    sh.column_dimensions["H"].width = 25
    sh.column_dimensions["I"].width = 20

    book.save(f"teams{actual_date[0]}_{actual_date[1]}_{actual_date[2]}.xlsx")

def add_teams(arr_teams, actual_date):
    book = openpyxl.load_workbook(filename=f"teams{actual_date[0]}_{actual_date[1]}_{actual_date[2]}.xlsx")
    sh = book.active
    row = 1
    for r, team in enumerate(arr_teams, start=2):
        row+=1
        img = Image(team.logo)
        img.width = 65
        img.height = 65
        img.anchor = sh.cell(row, 1).coordinate
        sh.add_image(img)
        sh.row_dimensions[row].height = 50
        sh.cell(row,2).value = team.place
        sh.cell(row,3).value = team.points
        sh.cell(row,4).value = team.team_name
        sh.cell(row,4).alignment = Alignment(horizontal='center', vertical='center')
        sh.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        sh.cell(row,9).value = team.url
        sh.cell(row,9).hyperlink = f"https://www.hltv.org{team.url}"
        sh.cell(row,9).style = "Hyperlink"

        sh.cell(row+1,1).value = "Nickname"
        sh.cell(row+2,1).value = "Name"
        sh.cell(row+3,1).value = "Age"
        sh.cell(row+4,1).value = "Country"
        sh.cell(row+5,1).value = "Url"

        for column, player in enumerate(team.players, start=4):
            sh.cell(row+1,column).value = player.nickname
            sh.cell(row+2,column).value = player.name
            sh.cell(row+3,column).value = player.age
            sh.cell(row+4,column).value = player.country
            sh.cell(row+5,column).value = player.url
            sh.cell(row+5,column).hyperlink = f"https://www.hltv.org{player.url}"
            sh.cell(row+5,column).style = "Hyperlink"
        row+=5
            

    book.save(f"teams{actual_date[0]}_{actual_date[1]}_{actual_date[2]}.xlsx")
